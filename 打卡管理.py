# !/user/bin/env Python3
# -*- coding:utf-8 -*-
 
"""
file：window.py
create time:2021/10/29 15:34
author:Yiqing Liu
desc: 人事统计
install:python3 setup.py py2app -A
"""
import tkinter as tk
import tkinter.messagebox 
from tkinter import filedialog, dialog
import os
import csv
import datetime
import getopt
import sys
import os
import time
import pandas as pd
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment, PatternFill


class personalInfo:
    def createPersonalInfo(self, name, department):
        self.PersonalInfo = {
            "name": name,
            "department":department,
            'attandenceDay': 0,
            'businessLeaveHour': 0,
            'sickLeave': 0,
            'annualLeave': 0,
            'marryLeave': 0,
            'funeralLeave': 0,
            'maternityLeave': 0,
            'fullAttandence': False,
            'lateDay': 0,
            'earlyLeaveDay': 0,
            'overtimeWorkDayRegular': 0,  # 2=<overwork<3 Monday-Friday
            'overtimeWeekendRegulr': 0,  # 5=<work<9 weekend
            'overtimeWeekendExtra': 0,  # worktime>9 weekend
            'ovtimeSalary': 0,
            'missAttandence': 0,
            'autoAttendence': 0
        }
        self.newComer = False
        self.fullAttandenceDay = 0
        self.autoAttendenceMorning = 0
    def get(self, attri):
        return self.PersonalInfo[attri]

    def alter(self, attri, value):
        self.PersonalInfo[attri] = value

    def calculateOvtimeSalary(self):
        self.PersonalInfo["ovtimeSalary"] = self.PersonalInfo['overtimeWorkDayRegular'] * 50 + \
            self.PersonalInfo['overtimeWeekendRegulr'] * 200 + \
            self.PersonalInfo['overtimeWeekendExtra'] * 500

    def exportString(self):
        exportString = ""
        for item in self.PersonalInfo:
            exportString += str(self.PersonalInfo[item]) + ','
        exportString += '\n'

        return exportString

    def StringToMinute(self, stringIn):
        workTime = int(stringIn[:2])*60 + int(stringIn[3:])
        if(int(stringIn[:2]) < 5):
            workTime +=(24*60+int(stringIn[:2]) * 60)
        return workTime

    def evalLeave(self, leaveReason, leaveTime):
        if leaveReason == "事假":
            leaveHour = float(leaveTime[:-2])
            self.PersonalInfo['businessLeaveHour'] += leaveHour
        elif leaveReason == "病假":
            self.PersonalInfo['sickLeave'] += float(leaveTime[:-1])
        elif leaveReason == "婚假":
            self.PersonalInfo['marryLeave'] += 1
        elif leaveReason == "丧假":
            self.PersonalInfo['funeralLeave'] += 1
        elif leaveReason == "陪产假":
            self.PersonalInfo['maternityLeave'] += 1
        elif leaveReason == "年假":
            self.PersonalInfo['annualLeave'] += float(leaveTime[:-1])

    def evalFullAttandence(self, checkInTime, workDay, iswork_m):
        if self.StringToMinute(checkInTime) <= 540 and workDay and iswork_m:
            self.fullAttandenceDay += 1

    def fullAttandenceAward(self):
        if self.fullAttandenceDay+self.autoAttendenceMorning >= self.PersonalInfo["attandenceDay"] and not self.newComer:
            self.PersonalInfo["fullAttandence"]=True

    def evalCheckInResult(self, checkInResult, workday):
        if checkInResult == "迟到" and workday:
            self.PersonalInfo['lateDay'] += 1
        if checkInResult == "正常(当天新加入考勤组，系统自动打卡)":
            self.newComer = True

    def evalEarlyLeave(self, leaveResult, workday):
        if leaveResult == "早退" and workday:
            self.PersonalInfo['earlyLeaveDay'] += 1

    def evalOverwork(self, workDay, iswork, checkInTime, LeaveTime):
        try:
            workTime = self.StringToMinute(LeaveTime) - self.StringToMinute(checkInTime)
        except BaseException:
            workTime = 0
        if workDay :
            if iswork and workTime > 0:
                overtime = int((workTime - 540)/30) * 0.5
                self.PersonalInfo['overtimeWorkDayRegular'] += overtime
            
        else:
            if 300 <= workTime and workTime < 540:
                self.PersonalInfo['overtimeWeekendRegulr'] +=1
            elif workTime >= 540:
                self.PersonalInfo['overtimeWeekendExtra'] +=1

    def checkInExceptions(self, checkinResult, checkoutReslut,leave,workDay):
        if workDay and ("无需打卡" in checkinResult or "无需打卡" in checkoutReslut):
            self.PersonalInfo["autoAttendence"] +=1
            if ("请假" in checkinResult or "请假" in checkoutReslut ):
                self.PersonalInfo["autoAttendence"] -=1
        if workDay and ("无需打卡" in checkinResult):
            self.autoAttendenceMorning += 1
        elif ('缺卡' in checkinResult or '缺卡' in checkoutReslut):
            self.PersonalInfo['missAttandence'] +=1

    def evalSingleLine(self, line):
        workDay = False
        iswork = True
        iswork_m = True
        if line[7] == "早9晚6 09:00-18:00;可晚到1小时;可早走1小时":
            self.PersonalInfo['attandenceDay'] += 1
            if "无需打卡" in line[12] or "无需打卡" in line[9]:
                iswork = False
            if "无需打卡" in line[0]:
                iswork_m=false
            workDay = True
        self.evalLeave(line[20], line[19])
        try:
            self.evalFullAttandence(line[8],workDay,iswork_m)
        except BaseException:
            pass

        self.evalCheckInResult(line[9], workDay)
        self.evalEarlyLeave(line[12], workDay)
        self.evalOverwork(workDay,iswork,line[8],line[11])
        self.checkInExceptions(line[9],line[12],line[20], workDay)

 
window = tk.Tk()
window.title('打卡管理   by:617') # 标题
window.geometry('500x550') # 窗口尺寸


file_path = ''
 
file_text = ''
 
directory = ''
im = Image.open('tt.png')
tkimage=ImageTk.PhotoImage(image=im)
tbl=tk.Label(image=tkimage)
tbl.pack()

def save_excel(directory,title,checkInList):
    color = ['99b3e5']
    wb = Workbook()
    ws = wb.active
    ws.row_dimensions[1].height=40
    ws.title = '涛思打卡情况'
    ws.append(title.split(","))
    fille = PatternFill('solid', fgColor=color[0]) 
    ws.column_dimensions['A'].width=20
    ws.column_dimensions['B'].width=35
    for i in range(1,len(title.split(","))+1):
        ws.cell(row=1, column=i).alignment = Alignment(wrapText=True)
        ws.cell(row=1, column=i).fill = fille
        if i >= 3:
            ws.column_dimensions[chr(i -1 + ord('A'))].width=9.27
    for key in checkInList.keys():
        checkInList[key].fullAttandenceAward()
        checkInList[key].calculateOvtimeSalary()
        result = list(checkInList[key].PersonalInfo.values())
        ws.append(result)
    
    wb.save(directory+"/output.xlsx")
    

 
def open_file():
  '''
  打开文件
  :return:
  '''
  global file_path
  global file_text
  global directory
  file_path = filedialog.askopenfilename(title=u'选择文件', initialdir=(os.path.expanduser('~/Desktop')))
  directory = os.path.split(file_path)[0]
  #print('打开文件：', file_path)
  if file_path is not None:
    csvFileName = "".join([file_path[:-4],'csv'])
    read_file = pd.read_excel (file_path)
    read_file.to_csv (csvFileName, index = None, header=0, encoding='utf-8')

    # os.system(f'cp {fileName} {csvFileName}')
    checkInList = {}
    info = pd.read_csv(csvFileName)
    lineInfo = info.iloc[0]
    for i in range(len(info)):
        lineInfo = info.iloc[i]
        if lineInfo[0] not in checkInList:
            personalInfo
            newPerson = personalInfo()
            newPerson.createPersonalInfo(lineInfo[0],lineInfo[1])
            checkInList[lineInfo[0]]= newPerson
        checkInList[lineInfo[0]].evalSingleLine(lineInfo)
    # openResult = open(directory+'/output.csv','w',encoding='utf-8')
    title = "姓名,部门,本月考勤天数,事假(小时),病假（天）,年假（天）,婚假（天）,丧假（天）,陪产假（天）,满勤奖,迟到,早退,工作日加班总时长,5≤加班时长＜9,9≤加班时长,加班奖金,缺卡,无需打卡"
    save_excel(directory,title,checkInList)
    
    
    tkinter.messagebox.showinfo('提示','计算完成,文件在'+directory)
 

bt1 = tk.Button(window, text='打开文件', width=15, height=2, command=open_file)
bt1.pack()


window.mainloop() # 显示
